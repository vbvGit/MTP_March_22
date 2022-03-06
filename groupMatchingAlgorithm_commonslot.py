from openpyxl import Workbook,load_workbook
from collections import defaultdict
import pandas as pd
import time
import math
import re
import copy

'''
Future aspect : Apply dynamic behaviour to both task and applicants.
'''

'''
Data conversion functions from .xlsx to python dictionary
'''
# Convert the skills cell to list of skills and trim the extra spaces from each skill(assumes skill data in lowercase)
def preprocess_skills_xl(skill_str):
    skills = skill_str.split(',')
    fin_skill = []
    for i in range(len(skills)):
        s = re.sub(' +',' ',skills[i].strip())
        fin_skill.append(s)
    return fin_skill

# Convert the location and slot cells of excel i.e x_coord,y_coord and slot_start,slot_end
def preprocess_loc_slot_xl(x,y):
    return [int(x),int(y)]

# Convert thr Tasks.xlsx data row by row to Task dictionary 
def preprocessTaskData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    T = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row =98,min_col = 1,max_col = 4,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        T[row[0]].append(skills)
        T[row[0]].append(loacation)
    return dict(T)

# Convert thr MainTaskInfo.xlsx data row by row to MainTask dictionary 
def preprocess_MainTaskInfo(filename):
    wb = load_workbook(filename)
    ws = wb.active

    T = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row =98,min_col = 1,max_col = 2,values_only = True):
        T[row[0]].append(row[1])
    return dict(T)

# Convert Volunteer data row by row to Applicants dictionary
def preprocessVolunteerData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    A = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row = 1576,min_col = 1,max_col = 6,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        slot = preprocess_loc_slot_xl(row[4],row[5])
        A[row[0]].append(skills)
        A[row[0]].append(loacation)
        A[row[0]].append(slot)

    return dict(A)

# Preprocess the applicants data : O(AlogA)
def preprocessApplicatntSkills(A,reqSkills):
    for key,val in A.items():
        val[0][:] = [skill for skill in val[0] if reqSkills.get(skill,-1) != -1]

    A = dict(sorted(A.items(),key = lambda i : -len(i[1][0])))
    return A

def initVisistedTasks_Vw(Applicants):
    Vw = dict(zip(list(Applicants.keys()),[[] for i in range(len(Applicants))]))
    return Vw

def initResultantTeam_R_1(Tasks):
    R = {}
    for t,val in Tasks.items():
        R[t] = {}
        for skill in val[0]:
            R[t][skill] = [None,float('inf')]
    return R

def initResultantTeam_R_2(Tasks):
    R = {}
    for t,val in Tasks.items():
        R[t] = defaultdict(list)
    return R

def initAvailableApplicants(Applicants):
    availableApp = dict(zip(Applicants.keys(),[1 for i in range(len(Applicants))]))
    return availableApp

def computeCommonSkills(skillSetA,skillSetT):
    common = []
    for skill in skillSetA:
        if skill in skillSetT:
            common.append(skill)
    return common

def computeDist(locA,locB):
    return math.sqrt((locB[0]-locA[0])**2+(locB[1]-locA[1])**2)

def computeMinDistTask(locA,Tasks):
    minDistTask = None
    minDist = float('inf')
    for task,data in Tasks.items():
        dest = data[1]
        curDist = computeDist(locA,dest)
        if curDist<minDist:
            minDistTask = task
            minDist = curDist
    return minDistTask

# Drive VTM algorithm : O(A*T*Skills)
def groupMatchingAlgorithm_GMA(Tasks,Applicants,cost):

    visitedTasks_Vw = initVisistedTasks_Vw(Applicants)
    R_ResultantTeam_SkillsKey = initResultantTeam_R_1(Tasks)
    R_ResultantTeam_WorkersKey = initResultantTeam_R_2(Tasks)
    availableApplicants = initAvailableApplicants(Applicants)

    while True:
        # Gt = dict(zip(list(Tasks.keys()),[{} for i in range(len(Tasks.keys()))]))
        terminal = 1

        for applicant,availability in availableApplicants.items():
            if availableApplicants[applicant] == 1:
                # print()
                # print(f"Current available volunteer : {applicant}")
                minDistTask = computeMinDistTask(Applicants[applicant][1],Tasks)
                # print(f"Minimum distance task for volunteer {applicant} : {minDistTask}")
                if minDistTask not in visitedTasks_Vw[applicant]:
                    # Gt[minDistTask].add(applicant)
                    visitedTasks_Vw[applicant].append(minDistTask)

                    commonSkills = computeCommonSkills(Applicants[applicant][0],Tasks[minDistTask][0])
                    # print(f"Skills common between volunteer {applicant} and task : {minDistTask} : {commonSkills} ")
                    for skill in commonSkills:
                        curDist = computeDist(Tasks[minDistTask][1],Applicants[applicant][1])
                        if curDist < R_ResultantTeam_SkillsKey[minDistTask][skill][1]:
                            prevAssignedApplicant = R_ResultantTeam_SkillsKey[minDistTask][skill][0]
                            if prevAssignedApplicant is not None:
                                for s in R_ResultantTeam_WorkersKey[minDistTask][prevAssignedApplicant]:
                                    if R_ResultantTeam_SkillsKey[minDistTask][s][0] == prevAssignedApplicant:
                                        R_ResultantTeam_SkillsKey[minDistTask][s][0] = None
                                        R_ResultantTeam_SkillsKey[minDistTask][s][1] = float('inf')
                                del R_ResultantTeam_WorkersKey[minDistTask][prevAssignedApplicant]
                                availableApplicants[prevAssignedApplicant] = 1
                            R_ResultantTeam_SkillsKey[minDistTask][skill][0] = applicant
                            R_ResultantTeam_SkillsKey[minDistTask][skill][1] = curDist
                            R_ResultantTeam_WorkersKey[minDistTask][applicant].append(skill)
                            availableApplicants[applicant] = 0
                            terminal = 0
                    # print(f"R after assigning {applicant} to task {minDistTask}\n{R_ResultantTeam_SkillsKey}\n")
                # else:
                    # print(f"{minDistTask} is already visited by {applicant}")

        if terminal == 1 or availableApplicants == {}:
            break

    return R_ResultantTeam_SkillsKey


# Function to compute the Final slot for given threshold
def computeFinalSlot(time,thresh):
    # Stores the final slot with maximum time
    start = -1
    end = -1

    # Pointers to check the current maximum slot
    i = -1
    j = -1

    # To compare the next valid slot with the previous maximum
    curMaxTime = 0 

    # Loop through the time array in order to get the valid slot and update the final slot start and end if end-start+1 > curMaxTime
    while i<len(time):
        if time[i]>=thresh:
            j = i+1
            while j<len(time) and time[j]>=thresh:
                j+=1
            if j-i>curMaxTime:
                start = i
                end = j-1
                curMaxTime = j-i
            if j<len(time):
                i = j
            else:
                break
        else:
            i+=1

    # Re conver the start and end to its original form i.e MMMM -> HHMM
    if end != -1 and start != -1:
        if end-start<60:
            if start>60:
                start-=60
            if end<1379:
                end+=60
        s_hour = start//60
        s_minutes = start%60
        start = s_hour*100+s_minutes

        e_hour = end//60
        e_minutes = end%60
        end = e_hour*100+e_minutes

        return [start,end]
    return []

# Function to compute the threshold for freelancers
def getThreshFreelancers(n):
    # Initial threshold percentage = 100%
    thresh = 1

    # Set to store the number of freelancers from range [50% * n to 100% * n]
    n_thresh = set()

    # Compute the number of freelancers as per threshold eg 50% 10 = 5
    while thresh>=0.5:
        n_thresh.add(int(thresh*n))
        thresh-=0.1
    return n_thresh


# Function to preprocess the original time slots HHMM -> MMMM
def getPreprocessedSlots(slots):
    res = []
    for slot in slots:
        # Convert start time to minutes : HHMM -> 60*HH+MM
        s_hour = slot[0]//100
        s_minutes = slot[0]%100
        start = s_hour*60+s_minutes

        # Convert end time to minutes : HHMM -> 60*HH+MM
        e_hour = slot[1]//100
        e_minutes = slot[1]%100
        end = e_hour*60+e_minutes

        # Update preprocessed slots
        res.append([start,end])
    return res

# Function to compute the common time slot between all freelancers
def getCommonSlots(slots):
    # Preprocess the given slots in standard HHMM format to MMMM format
    preprocessed_slots = getPreprocessedSlots(slots)

    # Total minutes in a day can be from 0000 : 12AM to 1439 : 11:59PM
    time_slot_minutes = [0 for i in range(24*60+1)]

    # For each slot just mark the start marker i.e +=1 and the end marker i.e -=1 : T.C = O(len(slots))
    for slot in preprocessed_slots:
        # Mark the start time for given slot
        time_slot_minutes[slot[0]]+=1

        if slot[1]+1<len(time_slot_minutes):
            # Mark the end time for given slot
            time_slot_minutes[slot[1]+1]-=1


    # Prefix sum of the time_slot_minutes array in order to compute the count of slots in which that minute appeared.
    # Example if minute 15 -> 0015AM was a part of 10 queries then time_slot_minutes[15] = 10
    for i in range(1,len(time_slot_minutes)):
        time_slot_minutes[i]+=time_slot_minutes[i-1]

    # To store all the common slots found for the given number of threshold free lancers as key
    common_slots = {}

    # Number of threshold freelancers to look for [60% of total freelancers to 100% of total freelancers]
    n_thresh = getThreshFreelancers(len(slots))

    # Check if the common slot is available for the calculated thresholds
    for threshold in n_thresh:
        commonslot = computeFinalSlot(time_slot_minutes,threshold)

        # If available print and return
        if commonslot != []:
            # print(f"The common slot which is common among {threshold} freelancers is : {commonslot}.")
            common_slots[threshold] = commonslot
        # else:
        #     # If no slot available print
        #     # print(f"No common slot found which is common among {threshold} freelancers.")
    return common_slots




def findProposedSlot(slots):
    maxthresh = max(slots)
    return slots[maxthresh]


def computeSuccessRatio(R):
    completed = 0
    tasksCompleted = []
    Completed_R = {}
    for task,Skills in R.items():
        flag = 1
        for skill,info in Skills.items():
            if info[0] is None:
                flag = 0
                break
        if flag:
            completed+=1
            Completed_R[task] = copy.deepcopy(R[task])
            tasksCompleted.append(task)
    return (completed/len(R))*100,tasksCompleted,Completed_R

def computeNetUtilityScore(R,Applicants):
    utilityScoreDict = dict(zip(Applicants,[[0,0] for i in range(len(Applicants))]))
    NetUtilityScore = 0 
    for task,Skills in R.items():
        for skill,info in Skills.items():
            if info[0] is not None:
                utilityScoreDict[info[0]][0]+=1
                if utilityScoreDict[info[0]][1] == 0:
                    utilityScoreDict[info[0]][1] = info[1]
    utilityScores = dict(zip(Applicants,[0 for i in range(len(Applicants))]))
    for applicant,info in utilityScoreDict.items():
        if info[0] == 0:
            utilityScores[applicant] = 0
        elif info[1] == 0:
            utilityScores[applicant] = info[0]/3*info[0]
        else:
            utilityScores[applicant] = info[0]/info[1]
        NetUtilityScore+=utilityScores[applicant]
    return utilityScores,NetUtilityScore


def OverallCompletionRate_OCR(Task,R):
    totalCompletionRatio = 0
    for task,subtasks in Task.items():
        score  = 0 
        for subtask in subtasks:
            if subtask in R:
                score+=1
        totalCompletionRatio+= (score/len(subtasks))
    OCR = totalCompletionRatio/len(Task)
    return OCR


def computeSatisfactoryRate(slots,proposed_Slot):

    minute_slots = getPreprocessedSlots(slots)
    minute_proposed_Slot = getPreprocessedSlots([list(proposed_Slot)])[0]


    res = 0
    start2,end2 = minute_proposed_Slot[0],minute_proposed_Slot[1]
    # print(f"Start2 :{start2}\tEnd2 : {end2}")
    for start1,end1 in minute_slots:
        # print(f"Start1 :{start1}\tEnd1 : {end1}")
        temp = 0
        if start1>end2 or start2>end1:
            continue
        else:
            temp = (min(end2,end1)-max(start2,start1)+1)/(end2-start2+1)
            res+=temp
        # print(f"calculated satisfactoryRate : {temp}")
    return res



# Driver code for VTM and common slot
def driver(T_I,T,A,cost):
    # Map volunteers to tasks
    start = time.time()
    R = groupMatchingAlgorithm_GMA(T,A,1)
    successRatio_1,tasksCompleted,Final_R = computeSuccessRatio(R)
    completed = len(tasksCompleted)
    # print(completed)
    # print(Final_R)
    time_phase_1 = time.time()-start
    # print(completed)

    totalSatisfactoryRate_TSR = 0
    assignedApplicants = 0

    for task in tasksCompleted:
        volunteersMapped = set()
        for skill in R[task]:
            volunteersMapped.add(R[task][skill][0])

        volunteersMapped = list(volunteersMapped)

        slots = []
        for volunteer in volunteersMapped:
            slots.append(A[volunteer][2])
        commonslot = getCommonSlots(slots)
        if commonslot == {}:
            completed-=1
        else:
            proposed_Slot = findProposedSlot(commonslot)
            Final_R[task]["Slot"] = proposed_Slot
            satisfactoryRate = computeSatisfactoryRate(slots,proposed_Slot)
            totalSatisfactoryRate_TSR+=satisfactoryRate
            assignedApplicants+=len(slots)


    print("Tasks which got completed with slots are : ")
    print()
    print(Final_R)
    print()
    print()

    OCR = OverallCompletionRate_OCR(T_I,Final_R)

    totalSatisfactoryRate_TSR *= (assignedApplicants)/(math.sqrt((len(A)-assignedApplicants)**2+(assignedApplicants)**2))

    # print(completed)
    successRatio_2 = (completed/len(T))*100
    utilityScore,NetUtilityScore = computeNetUtilityScore(R,list(A.keys()))
    return R,successRatio_1,successRatio_2,utilityScore,NetUtilityScore,time_phase_1,OCR,totalSatisfactoryRate_TSR


start_time = time.time()
MainTaskInfo = preprocess_MainTaskInfo("MainTaskInfo.xlsx")
Tasks = preprocessTaskData_xl("Tasks.xlsx")
Applicants = preprocessVolunteerData_xl("Applicants.xlsx")
R,success_ratio_1,success_ratio_2,utilityScores,NetUtilityScore,time_phase_1,ocr,tsr = driver(MainTaskInfo,Tasks,Applicants,1)
print(f"\nFinal Result:\n\nR : {R}\n\nUtility scores for all participants:\n{utilityScores}\n\nSuccess_Ratio after Phase_1  = {success_ratio_1}\n\nSuccess_Ratio after Phase_2  = {success_ratio_2}\n\nNetUtilityScore = {NetUtilityScore}\n\nOverallCompletionRate_OCR = {ocr}\n\ntotalSatisfactoryRate_TSR = {tsr}\n\nTime taken to complete Phase_1 : {time_phase_1}\n\nTotal time taken : {time.time()-start_time}")



