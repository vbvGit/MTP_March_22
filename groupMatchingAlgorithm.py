from openpyxl import Workbook,load_workbook
from collections import defaultdict
import pandas as pd
import time
import math
import re

'''
Future aspect : Apply dynamic behaviour to both task and applicants.
'''

'''
Data conversion functions from .xlsx to python dictionary
'''
# Convert the skills cell to list of skills and trim the extra spaces from each skill(assumes skill data in lowercase)
def preprocess_skills_xl(skill_str):
    skills = skill_str.split(',')
    fin_skill = []
    for i in range(len(skills)):
        s = re.sub(' +',' ',skills[i].strip())
        fin_skill.append(s)
    return fin_skill

# Convert the location and slot cells of excel i.e x_coord,y_coord and slot_start,slot_end
def preprocess_loc_slot_xl(x,y):
    return [int(x),int(y)]

# Convert thr Tasks.xlsx data row by row to Task dictionary 
def preprocessTaskData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    T = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row =98,min_col = 1,max_col = 4,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        T[row[0]].append(skills)
        T[row[0]].append(loacation)
    return dict(T)

# Convert Volunteer data row by row to Applicants dictionary
def preprocessVolunteerData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    A = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row = 1576,min_col = 1,max_col = 6,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        slot = preprocess_loc_slot_xl(row[4],row[5])
        A[row[0]].append(skills)
        A[row[0]].append(loacation)
        A[row[0]].append(slot)

    return dict(A)

# Preprocess the applicants data : O(AlogA)
def preprocessApplicatntSkills(A,reqSkills):
    for key,val in A.items():
        val[0][:] = [skill for skill in val[0] if reqSkills.get(skill,-1) != -1]

    A = dict(sorted(A.items(),key = lambda i : -len(i[1][0])))
    return A

def initVisistedTasks_Vw(Applicants):
    Vw = dict(zip(list(Applicants.keys()),[[] for i in range(len(Applicants))]))
    return Vw

def initResultantTeam_R_1(Tasks):
    R = {}
    for t,val in Tasks.items():
        R[t] = {}
        for skill in val[0]:
            R[t][skill] = [None,float('inf')]
    return R

def initResultantTeam_R_2(Tasks):
    R = {}
    for t,val in Tasks.items():
        R[t] = defaultdict(list)
    return R

def initAvailableApplicants(Applicants):
    availableApp = dict(zip(Applicants.keys(),[1 for i in range(len(Applicants))]))
    return availableApp

def computeCommonSkills(skillSetA,skillSetT):
    common = []
    for skill in skillSetA:
        if skill in skillSetT:
            common.append(skill)
    return common

def computeDist(locA,locB):
    return math.sqrt((locB[0]-locA[0])**2+(locB[1]-locA[1])**2)

def computeMinDistTask(locA,Tasks):
    minDistTask = None
    minDist = float('inf')
    for task,data in Tasks.items():
        dest = data[1]
        curDist = computeDist(locA,dest)
        if curDist<minDist:
            minDistTask = task
            minDist = curDist
    return minDistTask

# Drive GMA algorithm : O(A*T*Skills)
def groupMatchingAlgorithm_GMA(Tasks,Applicants,cost):

    visitedTasks_Vw = initVisistedTasks_Vw(Applicants)
    R_ResultantTeam_SkillsKey = initResultantTeam_R_1(Tasks)
    R_ResultantTeam_WorkersKey = initResultantTeam_R_2(Tasks)
    availableApplicants = initAvailableApplicants(Applicants)

    while True:
        # Gt = dict(zip(list(Tasks.keys()),[{} for i in range(len(Tasks.keys()))]))
        terminal = 1

        for applicant,availability in availableApplicants.items():
            if availableApplicants[applicant] == 1:
                # print()
                # print(f"Current available volunteer : {applicant}")
                minDistTask = computeMinDistTask(Applicants[applicant][1],Tasks)
                # print(f"Minimum distance task for volunteer {applicant} : {minDistTask}")
                if minDistTask not in visitedTasks_Vw[applicant]:
                    # Gt[minDistTask].add(applicant)
                    visitedTasks_Vw[applicant].append(minDistTask)

                    commonSkills = computeCommonSkills(Applicants[applicant][0],Tasks[minDistTask][0])
                    # print(f"Skills common between volunteer {applicant} and task : {minDistTask} : {commonSkills} ")
                    for skill in commonSkills:
                        curDist = computeDist(Tasks[minDistTask][1],Applicants[applicant][1])
                        if curDist < R_ResultantTeam_SkillsKey[minDistTask][skill][1]:
                            prevAssignedApplicant = R_ResultantTeam_SkillsKey[minDistTask][skill][0]
                            if prevAssignedApplicant is not None:
                                for s in R_ResultantTeam_WorkersKey[minDistTask][prevAssignedApplicant]:
                                    if R_ResultantTeam_SkillsKey[minDistTask][s][0] == prevAssignedApplicant:
                                        R_ResultantTeam_SkillsKey[minDistTask][s][0] = None
                                        R_ResultantTeam_SkillsKey[minDistTask][s][1] = float('inf')
                                del R_ResultantTeam_WorkersKey[minDistTask][prevAssignedApplicant]
                                availableApplicants[prevAssignedApplicant] = 1
                            R_ResultantTeam_SkillsKey[minDistTask][skill][0] = applicant
                            R_ResultantTeam_SkillsKey[minDistTask][skill][1] = curDist
                            R_ResultantTeam_WorkersKey[minDistTask][applicant].append(skill)
                            availableApplicants[applicant] = 0
                            terminal = 0
                    # print(f"R after assigning {applicant} to task {minDistTask}\n{R_ResultantTeam_SkillsKey}\n")
                # else:
                    # print(f"{minDistTask} is already visited by {applicant}")

        if terminal == 1 or availableApplicants == {}:
            break

    return R_ResultantTeam_SkillsKey

def computeSuccessRatio(R):
    completed = 0
    tasksCompleted = []
    for task,Skills in R.items():
        flag = 1
        for skill,info in Skills.items():
            if info[0] is None:
                flag = 0
                break
        if flag:
            completed+=1
            tasksCompleted.append(task)
    return (completed/len(R))*100,tasksCompleted

def computeNetUtilityScore(R,Applicants):
    utilityScoreDict = dict(zip(Applicants,[[0,0] for i in range(len(Applicants))]))
    NetUtilityScore = 0 
    for task,Skills in R.items():
        for skill,info in Skills.items():
            if info[0] is not None:
                utilityScoreDict[info[0]][0]+=1
                if utilityScoreDict[info[0]][1] == 0:
                    utilityScoreDict[info[0]][1] = info[1]
    utilityScores = dict(zip(Applicants,[0 for i in range(len(Applicants))]))
    for applicant,info in utilityScoreDict.items():
        if info[0] == 0:
            utilityScores[applicant] = 0
        elif info[1] == 0:
            utilityScores[applicant] = info[0]/3*info[0]
        else:
            utilityScores[applicant] = info[0]/info[1]
        NetUtilityScore+=utilityScores[applicant]
    return utilityScores,NetUtilityScore


# Driver code for VTM and common slot
def driver(T,A,cost):
    # Map volunteers to tasks
    start = time.time()
    R = groupMatchingAlgorithm_GMA(T,A,1)
    successRatio_1,tasksCompleted = computeSuccessRatio(R)
    completed = len(tasksCompleted)
    time_phase_1 = time.time()-start
    # print(completed)
    for task in tasksCompleted:
        volunteersMapped = set()
        for skill in R[task]:
            volunteersMapped.add(R[task][skill][0])

        volunteersMapped = list(volunteersMapped)
        print(volunteersMapped)
        flag = 1
        # print()
        # print(R[task])
        # print(task)
        # print(A[volunteersMapped[0]][2])
        for volunteer in volunteersMapped[1:]:
            print(A[volunteer][2])
            if A[volunteer][2] != A[volunteersMapped[0]][2]:
                flag = 0
        if flag == 0:
            completed-=1
    # print(completed)
    successRatio_2 = (completed/len(T))*100
    utilityScore,NetUtilityScore = computeNetUtilityScore(R,list(A.keys()))
    return R,successRatio_1,successRatio_2,utilityScore,NetUtilityScore,time_phase_1


start_time = time.time()
Tasks = preprocessTaskData_xl("Tasks.xlsx")
Applicants = preprocessVolunteerData_xl("Applicants.xlsx")
R,success_ratio_1,success_ratio_2,utilityScores,NetUtilityScore,time_phase_1 = driver(Tasks,Applicants,1)
print(f"\nFinal Result:\n\nR : {R}\n\nUtility scores for all participants:\n{utilityScores}\n\nSuccess_Ratio after Phase_1  = {success_ratio_1}\n\nSuccess_Ratio after Phase_2  = {success_ratio_2}\n\nNetUtilityScore = {NetUtilityScore}\n\nTime taken to complete Phase_1 : {time_phase_1}\n\nTotal time taken : {time.time()-start_time}")