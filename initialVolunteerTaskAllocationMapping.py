from openpyxl import Workbook,load_workbook
from collections import defaultdict
import pandas as pd
import time
import math
import re
import copy

'''
Future aspect : Apply dynamic behaviour to both task and applicants.
'''

'''
Data conversion functions from .xlsx to python dictionary
'''
# Convert the skills cell to list of skills and trim the extra spaces from each skill(assumes skill data in lowercase)
def preprocess_skills_xl(skill_str):
    skills = skill_str.split(',')
    fin_skill = []
    for i in range(len(skills)):
        s = re.sub(' +',' ',skills[i].strip())
        fin_skill.append(s)
    return fin_skill

# Convert the location and slot cells of excel i.e x_coord,y_coord and slot_start,slot_end
def preprocess_loc_slot_xl(x,y):
    return [int(x),int(y)]

# Convert thr Tasks.xlsx data row by row to Task dictionary 
def preprocessTaskData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    T = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row = 98,min_col = 1,max_col = 4,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        T[row[0]].append(skills)
        T[row[0]].append(loacation)
    return dict(T)


# Convert thr MainTaskInfo.xlsx data row by row to MainTask dictionary 
def preprocess_MainTaskInfo(filename):
    wb = load_workbook(filename)
    ws = wb.active

    T = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row =98,min_col = 1,max_col = 2,values_only = True):
        T[row[0]].append(row[1])
    return dict(T)

# Convert Volunteer data row by row to Applicants dictionary
def preprocessVolunteerData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    A = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row = 1576,min_col = 1,max_col = 6,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        slot = preprocess_loc_slot_xl(row[4],row[5])
        A[row[0]].append(skills)
        A[row[0]].append(loacation)
        A[row[0]].append(slot)

    return dict(A)


'''
Unused Module
'''
# # Make subtask for each task location i.e task at each location is considered to be an individual task -- O(Total Task Locations : T)
# def preprocessTaskSubTasks(Tasks):
#     newTask = defaultdict(list)
#     for key,val in Tasks.items():
#         subTask = 1
#         for i in range(len(val[1])):
#             subTaskKey = key+str(subTask)
#             newTask[subTaskKey].append(val[0])
#             newTask[subTaskKey].append(val[1][i])
#             subTask+=1
#     return dict(newTask)




'''
Phase 1 functions : Volunteer task mapping algorithm
'''
# Generate G_ST -- O(T*Skills)
def generateSkillTaskMapperMatrix(T):
    # Map skill to G_ST matrix row indices
    skillIdx = {}
    row_idx = 0

    # Map task to G_ST matrix column indices
    taskIdx = {}
    col_idx = 0

    # First pass on task T to set the indices for each task and skill
    for key,val in T.items():
        taskIdx[key] = col_idx
        col_idx+=1
        for i in range(len(val[0])):
            if skillIdx.get(val[0][i],-1) == -1:
                skillIdx[val[0][i]] = row_idx
                row_idx+=1

    numOnes = 0
    taskCompletionInfo = dict(zip(list(taskIdx.keys()),[0 for i in range(len(T))]))
    # Second pass on task T to set the counts on the desired indices of G_ST matrix
    G_ST = [[0 for j in range(len(taskIdx))]for i in range(len(skillIdx))]
    for key,val in T.items():
        j = taskIdx[key]
        for skill in val[0]:
            i = skillIdx[skill]
            G_ST[i][j] = 1
            numOnes+=1
            taskCompletionInfo[key]+=1
    
    return G_ST,skillIdx,taskIdx,numOnes,taskCompletionInfo

# Preprocess the applicants data : O(AlogA)
def preprocessApplicatntSkills(A,reqSkills):
    for key,val in A.items():
        val[0][:] = [skill for skill in val[0] if reqSkills.get(skill,-1) != -1]

    A = dict(sorted(A.items(),key = lambda i : -len(i[1][0])))
    return A

# Returns the list of tasks which match the skillset of a particular applicant : O(T*Skills)
def recommendedTasks(volunteerData,skillIdx,taskIdx,G_ST):
    recommend = {}
    matchedSkills = dict(zip(taskIdx.keys(),[[] for i in range(len(taskIdx))]))
    maximumMatchedSkills = 0
    for task in taskIdx:
        j = taskIdx[task]
        for skill in volunteerData[0]:
            i = skillIdx[skill]
            if G_ST[i][j]>0:
                matchedSkills[task].append(skill)
                maximumMatchedSkills = max(len(matchedSkills[task]),maximumMatchedSkills)

    if maximumMatchedSkills == 0:
        return []

    for key,val in matchedSkills.items():
        if len(val) == maximumMatchedSkills:
            recommend[key] = val
    # print(f"Current G_ST : {G_ST}\nVolunteer data : {volunteerData}\nThe tasks recommended for this volunteer are : {recommend}")
    return recommend

# Compute the nearest task among all the given tasks for the given applicant location : O(T)
def findCostEfficientTask(volunteer_coordinate,validTasks,validTasksCoordinate,cost):
    # print(volunteer_coordinate)
    # print(validTasks)
    # print(validTasksCoordinate)
    minCost = float('inf')
    minCostTaskIdx = len(validTasks)
    for i in range(len(validTasks)):
        source = volunteer_coordinate
        destination = validTasksCoordinate[i]
        curDist = math.sqrt((destination[0]-source[0])**2+(destination[1]-source[1])**2)
        # print(f"Distance from volunteer to {validTasks[i]} : {curDist}")
        curCost = curDist*cost
        if curCost<minCost:
            minCost = curCost
            minCostTaskIdx = i
    return validTasks[minCostTaskIdx],minCost

# Updates the G_ST as per the skills fulfilled by the given applicant : O(Skills)
def updateGST(t_idx,skills,skillIdx,G_ST):
    skillsFulfilled = 0
    for skill in skills:
        s_idx = skillIdx[skill]
        G_ST[s_idx][t_idx]-=1
        skillsFulfilled+=1
    return skillsFulfilled

# Drive VTM algorithm : O(A*T*Skills)
def initialVolunteerTaskMappingAlgorithm(Tasks,Applicants,cost):

    # Make subtask for each task location i. task at each location is considered to be an individual task : O(T)
    # Tasks = preprocessTaskSubTasks(Tasks)

    '''
    Generate G_ST                                                                                         : O(T*Skills)
    skillIdx : Map each skill with the corresponding row index of G_ST
    taskIdx : Map each task with corresponding column matrix of G_ST
    numSkills : mber of 1's in G_ST which depicts the number of skills to be satisfied in total.
    '''
    G_ST,skillIdx,taskIdx,numSkills,taskCompletionInfo = generateSkillTaskMapperMatrix(Tasks)
    # print(G_ST)
    # print(skillIdx)
    # print(numSkills)
    # print(taskCompletionInfo)

    '''
    Preprocess the applicants data:                                                                       : O(AlogA)
    --> Delete all the skills which are not coming under required skills for each applicant.
    --> After deletion sort the applicants in decreasing order of the number of skills they have.
    '''
    Applicants = preprocessApplicatntSkills(Applicants,skillIdx)


    # The final map containing each task as key and the list of names of candidates assigned to that task as value
    volunteerTaskMap = defaultdict(list)

    # # The metric used to compute the final utility of all the volunteers
    # NetUtilityScore = 0

    # Iterate over the applicants to assign it to a particular task : O(A*(T*Skills + Skills + T)) = O(T*A*Skills)
    for volunteer,data in Applicants.items():
        # print(f"Current volunteer : {volunteer}")

        # Returns the list of tasks which match the skillset of a particular applicant
        suggestedTasks = recommendedTasks(data,skillIdx,taskIdx,G_ST)
        # print(suggestedTasks)

        # If no task matches the skillset of the applicant then go for the next applicant
        if len(suggestedTasks) == 0:
            continue
        else:
            '''
            If tasks are suggested for the given applicant then we need to compare the most optimal task 
            based on the distance of that applicant from each task and then assign it to the nearest task
            '''
            selectedtask,costIncured = findCostEfficientTask(data[1],list(suggestedTasks.keys()),[Tasks[t][1] for t in suggestedTasks.keys()],cost)
            
            # Update the G_ST
            t_idx = taskIdx[selectedtask]
            contributedSkills = updateGST(t_idx,suggestedTasks[selectedtask],skillIdx,G_ST)
            numSkills-=contributedSkills
            taskCompletionInfo[selectedtask]-=contributedSkills

            # # Compute the utility score for allocated volunteer
            # skillsCovered = len(set(Applicants[volunteer][0]).intersection(set(list(skillIdx.keys()))))
            # utilityScore = skillsCovered/costIncured
            # NetUtilityScore+=utilityScore

            volunteerTaskMap[selectedtask].append((volunteer,len(Applicants[volunteer][0]),costIncured))
            # print(f"GST after assigning volunteer : {volunteer} to task : {selectedtask} :\n{G_ST}\n\n")

        if numSkills == 0:
            return True,dict(volunteerTaskMap),G_ST,taskCompletionInfo

    return False,dict(volunteerTaskMap),G_ST,taskCompletionInfo


'''
Phase 2 functions : Common slot finding algorithm
'''
# Function to compute the Final slot for given threshold
def computeFinalSlot(time,thresh):
    # Stores the final slot with maximum time
    start = -1
    end = -1

    # Pointers to check the current maximum slot
    i = -1
    j = -1

    # To compare the next valid slot with the previous maximum
    curMaxTime = 0 

    # Loop through the time array in order to get the valid slot and update the final slot start and end if end-start+1 > curMaxTime
    while i<len(time):
        if time[i]>=thresh:
            j = i+1
            while j<len(time) and time[j]>=thresh:
                j+=1
            if j-i>curMaxTime:
                start = i
                end = j-1
                curMaxTime = j-i
            if j<len(time):
                i = j
            else:
                break
        else:
            i+=1

    # Re conver the start and end to its original form i.e MMMM -> HHMM
    if end != -1 and start != -1:
        if end-start<60:
            if start>60:
                start-=60
            if end<1379:
                end+=60
        s_hour = start//60
        s_minutes = start%60
        start = s_hour*100+s_minutes

        e_hour = end//60
        e_minutes = end%60
        end = e_hour*100+e_minutes

        return [start,end]
    return []

# Function to compute the threshold for freelancers
def getThreshFreelancers(n):
    # Initial threshold percentage = 100%
    thresh = 1

    # Set to store the number of freelancers from range [50% * n to 100% * n]
    n_thresh = set()

    # Compute the number of freelancers as per threshold eg 50% 10 = 5
    while thresh>=0.5:
        n_thresh.add(int(thresh*n))
        thresh-=0.1
    return n_thresh

# Function to preprocess the original time slots HHMM -> MMMM
def getPreprocessedSlots(slots):
    res = []
    for slot in slots:
        # Convert start time to minutes : HHMM -> 60*HH+MM
        s_hour = slot[0]//100
        s_minutes = slot[0]%100
        start = s_hour*60+s_minutes

        # Convert end time to minutes : HHMM -> 60*HH+MM
        e_hour = slot[1]//100
        e_minutes = slot[1]%100
        end = e_hour*60+e_minutes

        # Update preprocessed slots
        res.append([start,end])
    return res

# Function to compute the common time slot between all freelancers
def getCommonSlots(slots):

    # Preprocess the given slots in standard HHMM format to MMMM format
    preprocessed_slots = getPreprocessedSlots(slots)

    # Total minutes in a day can be from 0000 : 12AM to 1439 : 11:59PM
    time_slot_minutes = [0 for i in range(24*60+1)]

    # For each slot just mark the start marker i.e +=1 and the end marker i.e -=1 : T.C = O(len(slots))
    for slot in preprocessed_slots:
        # Mark the start time for given slot
        time_slot_minutes[slot[0]]+=1

        if slot[1]+1<len(time_slot_minutes):
            # Mark the end time for given slot
            time_slot_minutes[slot[1]+1]-=1


    # Prefix sum of the time_slot_minutes array in order to compute the count of slots in which that minute appeared.
    # Example if minute 15 -> 0015AM was a part of 10 queries then time_slot_minutes[15] = 10
    for i in range(1,len(time_slot_minutes)):
        time_slot_minutes[i]+=time_slot_minutes[i-1]

    # To store all the common slots found for the given number of threshold free lancers as key
    common_slots = {}

    # Number of threshold freelancers to look for [60% of total freelancers to 100% of total freelancers]
    n_thresh = getThreshFreelancers(len(slots))

    # Check if the common slot is available for the calculated thresholds
    for threshold in n_thresh:
        commonslot = computeFinalSlot(time_slot_minutes,threshold)

        # If available print and return
        if commonslot != []:
            # print(f"The common slot which is common among {threshold} freelancers is : {commonslot}.")
            common_slots[threshold] = commonslot
        # else:
        #     # If no slot available print
        #     # print(f"No common slot found which is common among {threshold} freelancers.")
    return common_slots



def findProposedSlot(slots):
    maxthresh = max(slots)
    return slots[maxthresh]




'''
Metric calculation functions : Comparison metrices
'''
# Function to compute the net utility score
def computeNetUtilityScore(VTM,Applicants):
    utilityScoreDict = dict(zip(Applicants,[0 for i in range(len(Applicants))]))
    NetUtilityScore = 0 
    for task,volunteers in VTM.items():
        for volunteerInfo in volunteers[:-1]:
            skillsMatched = volunteerInfo[1]
            costIncured = volunteerInfo[2]
            utilityScoreDict[volunteerInfo[0]]+=(skillsMatched/costIncured)
            NetUtilityScore+=utilityScoreDict[volunteerInfo[0]]
    return utilityScoreDict,NetUtilityScore 

def computeSuccessRatio_1(taskCompletionInfo):
    completed = 0
    for task,remSkills in taskCompletionInfo.items():
        if remSkills == 0:
            completed+=1
    return (completed/len(taskCompletionInfo))*100



def OverallCompletionRate_OCR(Task,VTM):
    totalCompletionRatio = 0
    for task,subtasks in Task.items():
        score  = 0 
        for subtask in subtasks:
            if subtask in VTM:
                score+=1
        totalCompletionRatio+= (score/len(subtasks))
    OCR = totalCompletionRatio/len(Task)
    return OCR


def computeSatisfactoryRate(slots,proposed_Slot):

    minute_slots = getPreprocessedSlots(slots)
    minute_proposed_Slot = getPreprocessedSlots([list(proposed_Slot)])[0]


    res = 0
    start2,end2 = minute_proposed_Slot[0],minute_proposed_Slot[1]
    # print(f"Start2 :{start2}\tEnd2 : {end2}")
    for start1,end1 in minute_slots:
        # print(f"Start1 :{start1}\tEnd1 : {end1}")
        temp = 0
        if start1>end2 or start2>end1:
            continue
        else:
            temp = (min(end2,end1)-max(start2,start1)+1)/(end2-start2+1)
            res+=temp
        # print(f"calculated satisfactoryRate : {temp}")
    return res


'''
Driver function
'''
# Driver code for VTM and common slot
def driver(T_I,T,A,cost):
    # Map volunteers to tasks
    start = time.time()
    completed,VTM,G_ST,taskCompletionInfo = initialVolunteerTaskMappingAlgorithm(T,A,1)
    success_ratio_1 = computeSuccessRatio_1(taskCompletionInfo)
    time_phase_1 = time.time()-start

    Final_VTM = {}

    totalCompleted = 0

    totalSatisfactoryRate_TSR = 0
    assignedApplicants = 0

    for task in taskCompletionInfo:
        if taskCompletionInfo[task] == 0:
            volunteersMapped = [volInfo[0] for volInfo in VTM[task]]
            slots = []
            for volunteer in volunteersMapped:
                slots.append(A[volunteer][2])
            commonslot = getCommonSlots(slots)
            if commonslot == {}:
                continue
            else:
                totalCompleted+=1
                proposed_Slot = findProposedSlot(commonslot)
                VTM[task].append(proposed_Slot)
                satisfactoryRate = computeSatisfactoryRate(slots,proposed_Slot)
                totalSatisfactoryRate_TSR+=satisfactoryRate
                assignedApplicants+=len(slots)
                Final_VTM[task] = copy.deepcopy(VTM[task])

    success_ratio_2 = (totalCompleted/len(taskCompletionInfo))*100
    utilityScoreDict,NetUtilityScore = computeNetUtilityScore(VTM,list(A.keys()))

    print("Tasks which got completed with slots are : ")
    print()
    print(Final_VTM)
    print()
    print()

    OCR = OverallCompletionRate_OCR(T_I,Final_VTM)

    totalSatisfactoryRate_TSR *= (assignedApplicants)/(math.sqrt((len(A)-assignedApplicants)**2+(assignedApplicants)**2))

    return VTM,success_ratio_1,success_ratio_2,utilityScoreDict,NetUtilityScore,time_phase_1,OCR,totalSatisfactoryRate_TSR


start_time = time.time()
MainTaskInfo = preprocess_MainTaskInfo("MainTaskInfo.xlsx")
Tasks = preprocessTaskData_xl("Tasks.xlsx")
Applicants = preprocessVolunteerData_xl("Applicants.xlsx")
VTM,success_ratio_1,success_ratio_2,utilityScores,NetUtilityScore,time_phase_1,ocr,tsr = driver(MainTaskInfo,Tasks,Applicants,1)
print(f"\nFinal Result:\n\nVTM : {VTM}\n\nUtility scores for all participants:\n{utilityScores}\n\nSuccess_Ratio after Phase_1  = {success_ratio_1}\n\nSuccess_Ratio after Phase_2  = {success_ratio_2}\n\nNetUtilityScore = {NetUtilityScore}\n\nNetUtilityScore = {NetUtilityScore}\n\nOverallCompletionRate =  {ocr}\n\nTotalSatisfactoryRate =  {tsr}\n\nTotal time taken : {time.time()-start_time}")

